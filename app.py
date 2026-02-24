import csv
import json
import math
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

CONFIG_PATH = Path.home() / ".excel_merge_ui_config.json"
VALID_EXTENSIONS = {".xlsx", ".csv"}
REQUIRED_COLUMNS = {"CHNumber", "TESTRESULT"}
SN_COLUMN_CANDIDATES = ("TESTSN", "SN")
SORTING_FIELDS = [
    {
        "label": "DDMI_Bias(mA)",
        "aliases": ["DDMI_Bias(mA)", "DDMI BIAS", "DDMI_BIAS"],
    },
    {"label": "Outer_OMA(dB)", "aliases": ["Outer_OMA(dB)", "Outer OMA"]},
    {"label": "Outer_ER(dB)", "aliases": ["Outer_ER(dB)", "Outer ER"]},
    {"label": "TDECQ(dB)", "aliases": ["TDECQ(dB)", "TDECQ"]},
    {"label": "RLM", "aliases": ["RLM"]},
    {"label": "Ceq(dB)", "aliases": ["Ceq(dB)", "Ceq", "CEQ"]},
    {
        "label": "TDECQ_Ceq(dB)",
        "aliases": ["TDECQ_Ceq(dB)", "TDECQ_Ceq", "TDECQ CEQ"],
    },
    {"label": "Overshoot", "aliases": ["Overshoot"]},
    {"label": "Undershoot", "aliases": ["Undershoot"]},
    {"label": "OMA_Sen_MSB", "aliases": ["OMA_Sen_MSB", "OMA Sen MSB"]},
    {"label": "OMA_Sen_LSB", "aliases": ["OMA_Sen_LSB", "OMA Sen LSB"]},
    {
        "label": "dTxP",
        "aliases": ["dTxP", "dTxP\\n", "dTxP(dB)", "dTxP (dB)", "DTXP"],
    },
    {"label": "dRxP1", "aliases": ["dRxP1", "dRxP"]},
]
PRIORITY_CHOICES = ["不啟用"] + [str(i) for i in range(1, 21)]
NON_NUMERIC_OUTPUT_COLUMNS = {"TESTSN", "CHNumber", "TESTRESULT", "TESTDATE"}


def load_ui_config() -> Dict[str, object]:
    default: Dict[str, object] = {
        "input_folder": "",
        "output_folder": "",
        "enable_sorting": False,
        "sorting_rows": [],
    }

    if not CONFIG_PATH.exists():
        return default

    try:
        raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return default

    if "input_folder" not in raw and "last_folder" in raw:
        raw["input_folder"] = raw.get("last_folder", "")

    config = dict(default)
    if isinstance(raw, dict):
        config.update(raw)
    return config


def save_ui_config(config: Dict[str, object]) -> None:
    CONFIG_PATH.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def infer_temp_tag(file_path: Path) -> str:
    stem_upper = file_path.stem.upper()
    for tag in ("RT", "LT", "HT"):
        if re.search(rf"(^|[_\-\s]){tag}([_\-\s]|$)", stem_upper):
            return tag
    return "UNKN"


def infer_temp_tag_from_chnumber(value: object) -> str:
    text = str(value).strip().upper()
    m = re.search(r"(?:^|[_\-\s])(RT|LT|HT)(?:[_\-\s]|$)", text)
    if m:
        return m.group(1)
    return "UNKN"


def normalize_ch_number(value: object) -> str:
    text = str(value).strip()
    m = re.search(r"(\d+)", text)
    if not m:
        return text
    return str(int(m.group(1)))


def find_sn_column(headers: List[str]) -> str:
    normalized_map = {str(h).strip().upper(): h for h in headers}
    for candidate in SN_COLUMN_CANDIDATES:
        if candidate in normalized_map:
            return normalized_map[candidate]
    return ""


def parse_testdate(value: object) -> Optional[datetime]:
    text = str(value).strip()
    if not text:
        return None

    try:
        num = float(text)
        if num > 59:
            return datetime(1899, 12, 30) + timedelta(days=num)
    except ValueError:
        pass

    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def channel_sort_key(ch_tag: str) -> Tuple[int, int]:
    m = re.match(r"^(\d+)_([A-Z]+)$", ch_tag)
    if not m:
        return (99, 99)

    ch, tag = int(m.group(1)), m.group(2)
    tag_order = {"RT": 0, "LT": 1, "HT": 2}
    return (tag_order.get(tag, 3), ch)


def read_csv_rows(file_path: Path) -> List[Dict[str, str]]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return []
        return [dict(row) for row in reader]


def read_xlsx_rows(file_path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "讀取 xlsx 需要 openpyxl，請先安裝: pip install openpyxl"
        ) from exc

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows: List[Dict[str, str]] = []
    for row in rows[1:]:
        record = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = row[i] if i < len(row) else ""
            record[header] = "" if value is None else str(value)
        data_rows.append(record)

    return data_rows


def read_table_rows(file_path: Path) -> List[Dict[str, str]]:
    if file_path.suffix.lower() == ".csv":
        return read_csv_rows(file_path)
    return read_xlsx_rows(file_path)


def validate_and_transform_file(file_path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    issues: List[str] = []

    try:
        rows = read_table_rows(file_path)
    except Exception as exc:
        return [], [f"{file_path.name}: 讀取失敗 ({exc})"]

    if not rows:
        return [], [f"{file_path.name}: 無資料"]

    headers = list(rows[0].keys())
    sn_key = find_sn_column(headers)
    if not sn_key:
        return [], [f"{file_path.name}: 缺少必要欄位 ['TESTSN' 或 'SN']"]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        return [], [f"{file_path.name}: 缺少必要欄位 {sorted(missing)}"]

    file_tag = infer_temp_tag(file_path)

    grouped: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        sn = str(row.get(sn_key, "")).strip()
        if not sn:
            issues.append(f"{file_path.name}: 發現空白 {sn_key}，已略過")
            continue

        normalized = dict(row)
        normalized["TESTSN"] = sn
        raw_ch_number = row.get("CHNumber", "")
        normalized["CHNumber"] = normalize_ch_number(raw_ch_number)
        normalized["TEMP_TAG"] = (
            file_tag
            if file_tag != "UNKN"
            else infer_temp_tag_from_chnumber(raw_ch_number)
        )
        normalized["TESTRESULT"] = str(row.get("TESTRESULT", "")).strip().upper()

        grouped.setdefault(sn, []).append(normalized)

    valid_rows: List[Dict[str, str]] = []
    expected = {str(i) for i in range(1, 9)}

    for sn, group in grouped.items():
        by_channel: Dict[str, List[Dict[str, str]]] = {}
        for g in group:
            channel = str(g["CHNumber"])
            by_channel.setdefault(channel, []).append(g)

        channel_set = set(by_channel.keys())
        if not expected.issubset(channel_set):
            missing_channels = sorted(expected - channel_set, key=int)
            issues.append(
                f"{file_path.name}: TESTSN={sn} 缺少 CHNumber={missing_channels}，已略過"
            )
            continue

        if any(ch not in expected for ch in channel_set):
            extra_channels = sorted(channel_set - expected)
            issues.append(
                f"{file_path.name}: TESTSN={sn} 發現非 1~8 的 CHNumber={extra_channels}，已略過"
            )
            continue

        selected_rows: List[Dict[str, str]] = []
        failed_channels: List[str] = []
        for ch in sorted(expected, key=int):
            ch_rows = by_channel[ch]
            pass_rows = [row for row in ch_rows if row["TESTRESULT"] == "PASS"]
            pass_row = None
            if pass_rows:
                pass_row = max(
                    pass_rows,
                    key=lambda row: (
                        parse_testdate(row.get("TESTDATE", "")) or datetime.min,
                    ),
                )
            if not pass_row:
                failed_channels.append(ch)
                continue
            selected_rows.append(pass_row)

        if failed_channels:
            issues.append(
                f"{file_path.name}: TESTSN={sn} CHNumber={failed_channels} 沒有 PASS，已略過"
            )
            continue

        for g in selected_rows:
            item = dict(g)
            item["CHNumber"] = f"{item['CHNumber']}_{item['TEMP_TAG']}"
            item.pop("TEMP_TAG", None)
            valid_rows.append(item)

    return valid_rows, issues


def write_merged_output(output_folder_path: Path, rows: List[Dict[str, str]]) -> Path:
    all_headers: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                all_headers.append(key)

    output_xlsx = output_folder_path / "merged_output.xlsx"
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "merged"
        ws.append(all_headers)
        for row in rows:
            ws.append(build_output_row(row, all_headers))
        wb.save(output_xlsx)
        return output_xlsx
    except ImportError:
        output_csv = output_folder_path / "merged_output.csv"
        with output_csv.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_headers)
            writer.writeheader()
            writer.writerows([{h: convert_output_value(h, row.get(h, "")) for h in all_headers} for row in rows])
        return output_csv


def process_folder(
    input_folder_path: Path,
    output_folder_path: Path,
    enable_sorting: bool = False,
    sorting_configs: Optional[List[Dict[str, object]]] = None,
) -> Tuple[Path, List[str], int, int, Optional[int]]:
    files = sorted(
        [
            p
            for p in input_folder_path.iterdir()
            if p.is_file() and p.suffix.lower() in VALID_EXTENSIONS
        ]
    )

    if not files:
        raise ValueError("資料夾內沒有 xlsx/csv 檔案")

    merged_rows: List[Dict[str, str]] = []
    messages: List[str] = []

    for file_path in files:
        transformed, issues = validate_and_transform_file(file_path)
        messages.extend(issues)
        merged_rows.extend(transformed)

    if not merged_rows:
        raise ValueError("沒有任何符合規則的資料可合併")

    count_by_sn: Dict[str, int] = {}
    for row in merged_rows:
        sn = row["TESTSN"]
        count_by_sn[sn] = count_by_sn.get(sn, 0) + 1

    qualified_sns = {sn for sn, cnt in count_by_sn.items() if cnt == 24}
    for sn, cnt in sorted(count_by_sn.items()):
        if cnt != 24:
            messages.append(f"提醒: 合併後 TESTSN={sn} 筆數為 {cnt}，非 24 筆")

    merged_rows = [row for row in merged_rows if row["TESTSN"] in qualified_sns]
    if not merged_rows:
        raise ValueError("沒有任何 TESTSN 符合 24 筆規則可輸出")
    qualified_24_sn_count = len(qualified_sns)

    merged_rows.sort(key=lambda row: (row["TESTSN"], channel_sort_key(row.get("CHNumber", ""))))

    output_path = write_merged_output(output_folder_path, merged_rows)
    if output_path.suffix.lower() == ".csv":
        messages.append("提醒: 未安裝 openpyxl，輸出改為 merged_output.csv")
        return (
            output_path,
            messages,
            len(merged_rows),
            qualified_24_sn_count,
            None,
        )

    sorting_qualified_sn_count: Optional[int] = None
    if enable_sorting:
        active_configs = sorting_configs or []
        if not active_configs:
            raise ValueError("啟用 sorting 時至少需要一個有效條件")

        sorting_rows, sorting_messages, sorting_steps = build_sorting_rows(
            merged_rows,
            active_configs,
        )
        messages.extend(sorting_messages)
        append_sorting_sheet(output_path, sorting_rows)
        append_sum_sheet(output_path, merged_rows, sorting_rows, sorting_steps)
        sorting_qualified_sn_count = len({row["TESTSN"] for row in sorting_rows})
        messages.append(f"sorting 工作表完成：符合條件的資料筆數 {len(sorting_rows)}")
        messages.append("sum 工作表完成：已彙整 merged、sorting 與篩選步驟")

    return (
        output_path,
        messages,
        len(merged_rows),
        qualified_24_sn_count,
        sorting_qualified_sn_count,
    )


def parse_float(value: object) -> Optional[float]:
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def convert_output_value(header: str, value: object) -> object:
    if value is None:
        return ""

    if header in NON_NUMERIC_OUTPUT_COLUMNS or not isinstance(value, str):
        return value

    text = value.strip()
    if not text:
        return ""

    if not re.fullmatch(r"[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?", text):
        return value

    number = float(text)
    if not math.isfinite(number):
        return value

    if re.fullmatch(r"[+-]?\d+", text):
        return int(text)
    return number


def build_output_row(row: Dict[str, str], headers: List[str]) -> List[object]:
    return [convert_output_value(h, row.get(h, "")) for h in headers]


def normalize_column_name(name: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(name).upper())


def make_column_lookup(row: Dict[str, str]) -> Dict[str, str]:
    return {normalize_column_name(k): k for k in row.keys()}


def get_value_by_aliases(row: Dict[str, str], aliases: List[str]) -> str:
    lookup = make_column_lookup(row)
    for alias in aliases:
        key = lookup.get(normalize_column_name(alias))
        if key:
            return row.get(key, "")
    return ""


def build_sorting_rows(
    rows: List[Dict[str, str]],
    sorting_configs: List[Dict[str, object]],
) -> Tuple[List[Dict[str, str]], List[str], List[Dict[str, object]]]:
    messages: List[str] = []
    steps: List[Dict[str, object]] = []
    grouped: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(row["TESTSN"], []).append(row)

    candidate_sns = set()
    for sn, sn_rows in grouped.items():
        if len(sn_rows) != 24:
            messages.append(f"sorting: TESTSN={sn} 筆數 {len(sn_rows)} 非 24，已略過")
            continue
        candidate_sns.add(sn)

    ordered_configs = sorted(sorting_configs, key=lambda item: int(item["priority"]))
    for config in ordered_configs:
        label = str(config["label"])
        aliases = list(config["aliases"])
        min_value = float(config["min"])
        max_value = float(config["max"])

        previous_candidate_sns = set(candidate_sns)
        before_count = len(previous_candidate_sns)

        kept_sns = set()
        for sn in sorted(candidate_sns):
            sn_rows = grouped[sn]
            all_in_range = True
            for row in sn_rows:
                current = parse_float(get_value_by_aliases(row, aliases))
                if current is None or current < min_value or current > max_value:
                    all_in_range = False
                    break
            if all_in_range:
                kept_sns.add(sn)

        candidate_sns = kept_sns
        removed_sns = sorted(
            [sn for sn in previous_candidate_sns if sn not in candidate_sns],
            key=lambda sn: str(sn),
        )
        steps.append(
            {
                "priority": config["priority"],
                "field": label,
                "range": f"[{min_value}, {max_value}]",
                "candidate_before": before_count,
                "qualified_sn": len(candidate_sns),
                "filtered_out": max(before_count - len(candidate_sns), 0),
                "qualified_sn_list": sorted(candidate_sns, key=lambda sn: str(sn)),
                "filtered_out_sn_list": removed_sns,
            }
        )

        messages.append(
            "sorting step P{priority} {field} {range_text}: {before} -> {after} "
            "(淘汰 {filtered})".format(
                priority=config["priority"],
                field=label,
                range_text=f"[{min_value}, {max_value}]",
                before=before_count,
                after=len(candidate_sns),
                filtered=max(before_count - len(candidate_sns), 0),
            )
        )
        if not candidate_sns:
            break

    sorting_rows = [row for row in rows if row["TESTSN"] in candidate_sns]
    sorting_rows.sort(key=lambda row: (row["TESTSN"], channel_sort_key(row.get("CHNumber", ""))))
    return sorting_rows, messages, steps


def append_sorting_sheet(
    output_path: Path,
    sorting_rows: List[Dict[str, str]],
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    if "sorting" in wb.sheetnames:
        del wb["sorting"]

    all_headers: List[str] = []
    seen = set()
    for row in sorting_rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                all_headers.append(key)

    ws = wb.create_sheet("sorting")
    if all_headers:
        ws.append(all_headers)
        for row in sorting_rows:
            ws.append(build_output_row(row, all_headers))

    wb.save(output_path)


def append_sum_sheet(
    output_path: Path,
    merged_rows: List[Dict[str, str]],
    sorting_rows: List[Dict[str, str]],
    sorting_steps: List[Dict[str, object]],
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    if "sum" in wb.sheetnames:
        del wb["sum"]

    merged_sns = sorted(
        {row["TESTSN"] for row in merged_rows},
        key=lambda sn: str(sn),
    )
    sorting_sns = sorted(
        {row["TESTSN"] for row in sorting_rows},
        key=lambda sn: str(sn),
    )

    ws = wb.create_sheet("sum")
    ws.append(["Item", "Value"])
    ws.append(["Merged T", len(merged_sns)])
    ws.append(["Sorting T", len(sorting_sns)])
    for idx, step in enumerate(sorting_steps, start=1):
        ws.append(
            [
                f"Sorting Step {idx} T (P{step.get('priority')})",
                step.get("qualified_sn", 0),
            ]
        )
    ws.append([])

    stage_columns: List[Tuple[str, List[str]]] = [
        ("Merged TESTSN (24 rows each)", merged_sns),
        ("Sorting T", sorting_sns),
    ]
    for step in sorting_steps:
        stage_columns.append(
            (
                f"P{step.get('priority')} {step.get('field')} {step.get('range')}",
                list(step.get("qualified_sn_list", [])),
            )
        )

    ws.append([header for header, _ in stage_columns])
    max_len = max(len(values) for _, values in stage_columns)
    for idx in range(max_len):
        ws.append([values[idx] if idx < len(values) else "" for _, values in stage_columns])

    ws.append([])
    ws.append(
        [
            "Sorting Logic",
            "Range",
            "Priority",
            "Candidate SN before step",
            "Qualified SN after step",
            "Filtered out in this step",
            "Qualified TESTSN list",
            "Filtered out TESTSN list",
        ]
    )
    for step in sorting_steps:
        ws.append(
            [
                step["field"],
                step["range"],
                step["priority"],
                step.get("candidate_before", 0),
                step["qualified_sn"],
                step.get("filtered_out", 0),
                ", ".join(str(sn) for sn in step.get("qualified_sn_list", [])),
                ", ".join(str(sn) for sn in step.get("filtered_out_sn_list", [])),
            ]
        )

    wb.save(output_path)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 合併工具")
        self.geometry("760x500")

        self.ui_config = load_ui_config()
        self.folder_var = tk.StringVar(value=str(self.ui_config.get("input_folder", "")))
        self.output_folder_var = tk.StringVar(
            value=str(self.ui_config.get("output_folder", ""))
        )
        self.enable_sorting_var = tk.BooleanVar(
            value=bool(self.ui_config.get("enable_sorting", False))
        )
        self.sorting_rows_vars: List[Dict[str, tk.StringVar]] = []
        self._build_ui()
        self._apply_saved_sorting_rows()

    def _build_ui(self):
        frame = ttk.Frame(self, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="資料夾路徑：").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.folder_var, width=80).grid(
            row=1, column=0, columnspan=2, sticky="ew", padx=(0, 8), pady=(4, 8)
        )
        ttk.Button(frame, text="選擇資料夾", command=self.choose_folder).grid(
            row=1, column=2, sticky="ew"
        )

        ttk.Label(frame, text="輸出檔案資料夾：").grid(row=2, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.output_folder_var, width=80).grid(
            row=3, column=0, columnspan=2, sticky="ew", padx=(0, 8), pady=(4, 8)
        )
        ttk.Button(frame, text="選擇資料夾", command=self.choose_output_folder).grid(
            row=3, column=2, sticky="ew"
        )

        ttk.Checkbutton(
            frame,
            text="啟用 sorting（可設定多個條件優先順序）",
            variable=self.enable_sorting_var,
        ).grid(row=4, column=0, columnspan=3, sticky="w", pady=(0, 4))

        ttk.Label(frame, text="Sorting 條件（Min/Max + 優先順序 1~20）：").grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(0, 4)
        )

        sorting_frame = ttk.Frame(frame)
        sorting_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(0, 6))
        sorting_frame.columnconfigure(0, weight=2)
        sorting_frame.columnconfigure(1, weight=1)
        sorting_frame.columnconfigure(2, weight=1)
        sorting_frame.columnconfigure(3, weight=1)

        ttk.Label(sorting_frame, text="欄位").grid(row=0, column=0, sticky="w")
        ttk.Label(sorting_frame, text="Min").grid(row=0, column=1, sticky="w")
        ttk.Label(sorting_frame, text="Max").grid(row=0, column=2, sticky="w")
        ttk.Label(sorting_frame, text="Priority").grid(row=0, column=3, sticky="w")

        for idx, field in enumerate(SORTING_FIELDS, start=1):
            min_var = tk.StringVar()
            max_var = tk.StringVar()
            priority_var = tk.StringVar(value=PRIORITY_CHOICES[0])
            self.sorting_rows_vars.append(
                {
                    "label": field["label"],
                    "min_var": min_var,
                    "max_var": max_var,
                    "priority_var": priority_var,
                }
            )

            ttk.Label(sorting_frame, text=field["label"]).grid(row=idx, column=0, sticky="w")
            ttk.Entry(sorting_frame, textvariable=min_var, width=12).grid(
                row=idx, column=1, sticky="ew", padx=(0, 4)
            )
            ttk.Entry(sorting_frame, textvariable=max_var, width=12).grid(
                row=idx, column=2, sticky="ew", padx=(0, 4)
            )
            ttk.Combobox(
                sorting_frame,
                textvariable=priority_var,
                values=PRIORITY_CHOICES,
                state="readonly",
                width=8,
            ).grid(row=idx, column=3, sticky="w")

        ttk.Button(frame, text="執行", command=self.run_process).grid(
            row=7, column=0, columnspan=3, sticky="ew", pady=(0, 10)
        )

        ttk.Label(frame, text="執行訊息：").grid(row=8, column=0, sticky="w")

        self.log_text = tk.Text(frame, wrap="word", height=20)
        self.log_text.grid(row=9, column=0, columnspan=3, sticky="nsew")

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=9, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)

        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        frame.columnconfigure(2, weight=1)
        frame.rowconfigure(9, weight=1)

    def choose_folder(self):
        folder = filedialog.askdirectory(initialdir=self.folder_var.get() or ".")
        if folder:
            self.folder_var.set(folder)

    def choose_output_folder(self):
        folder = filedialog.askdirectory(initialdir=self.output_folder_var.get() or ".")
        if folder:
            self.output_folder_var.set(folder)

    def _collect_ui_config(self) -> Dict[str, object]:
        sorting_rows: List[Dict[str, str]] = []
        for row in self.sorting_rows_vars:
            sorting_rows.append(
                {
                    "label": row["label"],
                    "min": row["min_var"].get().strip(),
                    "max": row["max_var"].get().strip(),
                    "priority": row["priority_var"].get().strip(),
                }
            )

        return {
            "input_folder": self.folder_var.get().strip(),
            "output_folder": self.output_folder_var.get().strip(),
            "enable_sorting": self.enable_sorting_var.get(),
            "sorting_rows": sorting_rows,
        }

    def _save_ui_config(self) -> None:
        save_ui_config(self._collect_ui_config())

    def _apply_saved_sorting_rows(self) -> None:
        saved_rows = self.ui_config.get("sorting_rows", [])
        if not isinstance(saved_rows, list):
            return

        saved_by_label = {
            str(item.get("label", "")): item
            for item in saved_rows
            if isinstance(item, dict)
        }
        for row in self.sorting_rows_vars:
            saved = saved_by_label.get(row["label"])
            if not saved:
                continue

            row["min_var"].set(str(saved.get("min", "")))
            row["max_var"].set(str(saved.get("max", "")))
            priority = str(saved.get("priority", PRIORITY_CHOICES[0]))
            if priority not in PRIORITY_CHOICES:
                priority = PRIORITY_CHOICES[0]
            row["priority_var"].set(priority)

    def log(self, text: str):
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")

    def run_process(self):
        self.log_text.delete("1.0", "end")
        folder = self.folder_var.get().strip()
        if not folder:
            messagebox.showerror("錯誤", "請先選擇資料夾")
            return

        output_folder = self.output_folder_var.get().strip() or folder

        folder_path = Path(folder)
        if not folder_path.exists() or not folder_path.is_dir():
            messagebox.showerror("錯誤", "資料夾路徑不存在")
            return

        output_folder_path = Path(output_folder)
        if not output_folder_path.exists() or not output_folder_path.is_dir():
            messagebox.showerror("錯誤", "輸出資料夾路徑不存在")
            return

        self._save_ui_config()

        enable_sorting = self.enable_sorting_var.get()
        sorting_configs: List[Dict[str, object]] = []

        if enable_sorting:
            used_priorities = set()
            for field, row_vars in zip(SORTING_FIELDS, self.sorting_rows_vars):
                priority = row_vars["priority_var"].get().strip()
                if priority == PRIORITY_CHOICES[0]:
                    continue

                min_value = parse_float(row_vars["min_var"].get())
                max_value = parse_float(row_vars["max_var"].get())
                if min_value is None or max_value is None:
                    messagebox.showerror("錯誤", f"{field['label']} 需要有效的最小值與最大值")
                    return
                if min_value > max_value:
                    messagebox.showerror("錯誤", f"{field['label']} 最小值不可大於最大值")
                    return
                if priority in used_priorities:
                    messagebox.showerror("錯誤", f"優先順序 {priority} 重複，請調整")
                    return
                used_priorities.add(priority)
                sorting_configs.append(
                    {
                        "label": field["label"],
                        "aliases": field["aliases"],
                        "min": min_value,
                        "max": max_value,
                        "priority": priority,
                    }
                )

            if not sorting_configs:
                messagebox.showerror("錯誤", "啟用 sorting 時請至少設定一個條件")
                return

        try:
            output_path, messages, total_rows, qualified_24_sn_count, sorting_qualified_sn_count = process_folder(
                folder_path,
                output_folder_path,
                enable_sorting=enable_sorting,
                sorting_configs=sorting_configs,
            )
            self.log(f"完成，總筆數：{total_rows}")
            self.log(f"輸出檔案：{output_path}")
            self.log(f"符合 24 筆的 TESTSN 數量：{qualified_24_sn_count}")
            if enable_sorting:
                self.log(f"sorting 最後符合條件的 TESTSN 數量：{sorting_qualified_sn_count or 0}")
            for msg in messages:
                self.log(msg)
            messagebox.showinfo(
                "完成",
                self._build_completion_message(
                    output_path=output_path,
                    enable_sorting=enable_sorting,
                    sorting_configs=sorting_configs,
                    qualified_24_sn_count=qualified_24_sn_count,
                    sorting_qualified_sn_count=sorting_qualified_sn_count,
                ),
            )
        except Exception as exc:
            self.log(f"執行失敗：{exc}")
            messagebox.showerror("錯誤", str(exc))

    def _build_completion_message(
        self,
        output_path: Path,
        enable_sorting: bool,
        sorting_configs: List[Dict[str, object]],
        qualified_24_sn_count: int,
        sorting_qualified_sn_count: Optional[int],
    ) -> str:
        lines = [
            f"合併完成：{output_path}",
            f"符合 24 筆的 TESTSN 數量：{qualified_24_sn_count} 顆",
        ]
        if enable_sorting:
            lines.append(
                f"最後條件 sorting 後符合的 TESTSN 數量：{sorting_qualified_sn_count or 0} 顆"
            )
            lines.append("已選擇 sorting 條件：")
            for config in sorted(sorting_configs, key=lambda item: int(item["priority"])):
                lines.append(
                    "  - P{priority} {label}: Min={min_value}, Max={max_value}".format(
                        priority=config["priority"],
                        label=config["label"],
                        min_value=config["min"],
                        max_value=config["max"],
                    )
                )
        else:
            lines.append("最後條件 sorting 後符合的 TESTSN 數量：未啟用 sorting")
            lines.append("已選擇 sorting 條件：未啟用")

        return "\n".join(lines)


if __name__ == "__main__":
    app = App()
    app.mainloop()
